import io
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


CATEGORIES_CONFIG = {
    "Серверы": ["Процессор", "Оперативная память", "Накопитель", "Сетевой адаптер", "Блок питания", "Форм-актор"],
    "ПК": ["Процессор", "Оперативная память", "Накопитель", "Видеокарта", "Блок питания", "Форм-фактор"],
    "Мониторы": ["Диагональ", "Разрешение", "Тип матрицы", "Частота обновления", "Время отклика", "Интерфейсы"],
    "Моноблоки": ["Процессор", "Оперативная память", "Накопитель", "Диагональ", "Разрешение", "Видеокарта"],
    "Ноутбуки": ["Процессор", "Оперативная память", "Накопитель", "Диагональ", "Разрешение", "Видеокарта", "ОС"],
    "Планшеты": ["Процессор", "Оперативная память", "Накопитель", "Диагональ", "Разрешение", "ОС", "Связь"],
    "МФУ": ["Технология печати", "Формат", "Скорость печати", "Разрешение", "Сканирование", "Интерфейсы"],
    "Принтеры": ["Технология печати", "Формат", "Скорость печати", "Разрешение", "Интерфейсы"],
    "Периферия": ["Тип", "Интерфейс", "Совместимость"],
    "Сетевое оборудование": ["Тип", "Порты", "Скорость", "Протоколы", "Управление"],
    "Прочее": ["Описание", "Характеристики"]
}


def generate_template(dictionary: List[Dict], unitsByCategory: Dict = None, ktruValues: Dict = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Шаблон"
    
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws['A1'] = "Категория"
    ws['B1'] = "Наименование"
    ws['C1'] = "Код КТРУ"
    ws['D1'] = "Количество"
    
    for col in range(1, 5):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    col_idx = 5
    dict_by_category = {}
    for d in dictionary:
        cat = d.get('category')
        field = d.get('field_name')
        if cat and field:
            if cat not in dict_by_category:
                dict_by_category[cat] = []
            dict_by_category[cat].append(field)
    
    all_fields = set()
    for fields in dict_by_category.values():
        all_fields.update(fields)
    
    for field in sorted(all_fields):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = field
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        col_idx += 1
    
    for col in range(1, col_idx):
        ws.column_dimensions[get_column_letter(col)].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def fill_excel(items: List[Dict], metadata: Dict, dictionary: List[Dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Заказ"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws['A1'] = "Заказчик:"
    ws['B1'] = metadata.get('customer_name', '')
    ws['A2'] = "ИНН:"
    ws['B2'] = metadata.get('customer_inn', '')
    ws['A3'] = "Дата:"
    ws['B3'] = metadata.get('doc_date', '')
    ws['A4'] = "ID:"
    ws['B4'] = metadata.get('processing_id', '')
    ws['A5'] = "Файл-источник:"
    ws['B5'] = metadata.get('source_file', '')
    
    for row in range(1, 6):
        ws.cell(row=row, column=1).font = Font(bold=True)
    
    start_row = 7
    headers = ["№", "Категория", "Наименование", "Код КТРУ", "Кол-во"]
    
    dict_by_category = {}
    for d in dictionary:
        cat = d.get('category')
        field = d.get('field_name')
        if cat and field:
            if cat not in dict_by_category:
                dict_by_category[cat] = []
            dict_by_category[cat].append(field)
    
    found_fields = set()
    for item in items:
        for char in item.get('characteristics', []) or []:
            name = char.get('name')
            if name:
                found_fields.add(name)

    ordered_fields = []
    for fields in dict_by_category.values():
        for field in fields:
            if any(field.lower() == f.lower() for f in found_fields):
                ordered_fields.append(field)

    extra_fields = [f for f in found_fields if not any(f.lower() == df.lower() for df in ordered_fields)]
    sorted_fields = ordered_fields + sorted(extra_fields)
    headers.extend(sorted_fields)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    for row_idx, item in enumerate(items, start_row + 1):
        ws.cell(row=row_idx, column=1, value=row_idx - start_row).border = thin_border
        ws.cell(row=row_idx, column=2, value=item.get('category', '')).border = thin_border
        ws.cell(row=row_idx, column=3, value=item.get('name', '')).border = thin_border
        ws.cell(row=row_idx, column=4, value=item.get('ktru_code', '')).border = thin_border
        ws.cell(row=row_idx, column=5, value=item.get('quantity', 1)).border = thin_border
        
        characteristics = {c['name']: c['value'] for c in item.get('characteristics', [])}
        
        for col_idx, field in enumerate(sorted_fields, 6):
            value = characteristics.get(field, '')
            ws.cell(row=row_idx, column=col_idx, value=value).border = thin_border
    
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_consolidated_report(orders: List[Dict], dictionary: List[Dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводный отчет"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["Дата", "Заказчик", "ИНН", "ID обработки", "Файл-источник", "Категория", "Товар", "Код КТРУ", "Кол-во"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row_idx = 2
    for order in orders:
        metadata = order.get('metadata', {})
        items = order.get('items', [])
        
        for item in items:
            ws.cell(row=row_idx, column=1, value=metadata.get('doc_date', '')).border = thin_border
            ws.cell(row=row_idx, column=2, value=metadata.get('customer_name', '')).border = thin_border
            ws.cell(row=row_idx, column=3, value=metadata.get('customer_inn', '')).border = thin_border
            ws.cell(row=row_idx, column=4, value=metadata.get('processing_id', '')).border = thin_border
            ws.cell(row=row_idx, column=5, value=metadata.get('source_file', '')).border = thin_border
            ws.cell(row=row_idx, column=6, value=item.get('category', '')).border = thin_border
            ws.cell(row=row_idx, column=7, value=item.get('name', '')).border = thin_border
            ws.cell(row=row_idx, column=8, value=item.get('ktru_code', '')).border = thin_border
            ws.cell(row=row_idx, column=9, value=item.get('quantity', 1)).border = thin_border
            row_idx += 1
    
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def parse_dictionary_from_excel(file_content: bytes) -> Dict[str, List[str]]:
    wb = load_workbook(io.BytesIO(file_content))
    ws = wb.active
    
    result = {}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            category = str(row[0]).strip()
            if category not in result:
                result[category] = []
            
            for cell in row[1:]:
                if cell and str(cell).strip():
                    result[category].append(str(cell).strip())
    
    return result


def export_ktru_lookup_to_excel(ktru_code: str, data: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "КТРУ"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    
    ws['A1'] = "Код КТРУ:"
    ws['B1'] = ktru_code
    ws['A1'].font = Font(bold=True)
    
    ws['A3'] = "Данные"
    ws['A3'].font = header_font
    ws['A3'].fill = header_fill
    
    ws['A4'] = data
    
    ws.column_dimensions['A'].width = 80
    ws.merge_cells('A4:A50')
    ws['A4'].alignment = Alignment(wrap_text=True, vertical='top')
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def generate_ktru_template_with_values(ktru_code: str, characteristics: List[Dict]) -> bytes:
    """
    Generate Excel template with characteristic names, units, and possible values.
    characteristics: List of {"name": str, "unit": str, "possible_values": str}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "КТРУ Шаблон"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    ws['A1'] = "Код КТРУ:"
    ws['B1'] = ktru_code
    ws['A1'].font = Font(bold=True)
    ws.merge_cells('A1:B1')
    
    headers = ["Наименование характеристики", "Ед. изм.", "Возможные значения"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    row_idx = 4
    for char in characteristics:
        ws.cell(row=row_idx, column=1, value=char.get('name', '')).border = thin_border
        ws.cell(row=row_idx, column=2, value=char.get('unit', '')).border = thin_border
        ws.cell(row=row_idx, column=3, value=char.get('possible_values', '')).border = thin_border
        ws.cell(row=row_idx, column=3).alignment = Alignment(wrap_text=True)
        row_idx += 1
    
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 100
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def export_scanned_codes_xlsx(codes: List[Dict]) -> bytes:
    """
    Экспортирует сканированные коды КТРУ в Excel файл.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Коды КТРУ"
    
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ["Код КТРУ", "Группа", "Наименование", "Статус КТРУ", "Дата сканирования"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    included_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    excluded_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    row_idx = 2
    for code in codes:
        ktru_status = code.get('ktru_status') or code.get('status', '')
        
        ws.cell(row=row_idx, column=1, value=code.get('item_id', '')).border = thin_border
        ws.cell(row=row_idx, column=2, value=code.get('group_code', '')).border = thin_border
        ws.cell(row=row_idx, column=3, value=code.get('item_name', '')).border = thin_border
        
        status_cell = ws.cell(row=row_idx, column=4, value="Включено" if ktru_status == 'included' else ("Исключено" if ktru_status == 'excluded' else "—"))
        status_cell.border = thin_border
        if ktru_status == 'included':
            status_cell.fill = included_fill
        elif ktru_status == 'excluded':
            status_cell.fill = excluded_fill
        
        scanned_at = code.get('scanned_at')
        if scanned_at:
            try:
                dt = datetime.fromisoformat(scanned_at.replace('Z', '+00:00'))
                ws.cell(row=row_idx, column=5, value=dt.strftime('%d.%m.%Y %H:%M')).border = thin_border
            except:
                ws.cell(row=row_idx, column=5, value=scanned_at).border = thin_border
        else:
            ws.cell(row=row_idx, column=5, value='').border = thin_border
        
        row_idx += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    ws.auto_filter.ref = f"A1:E{row_idx-1}"
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
